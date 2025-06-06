import requests
import json
from typing import Dict, Optional

# API端点
API_URL = "https://devops-api.nullmax.net/asset/api/v1/fixed/transfer/reallocate"

# 认证令牌（实际使用时建议从环境变量或配置文件中获取）
AUTH_TOKEN = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJleHAiOjE3NTExNjI2NjIsImlhdCI6MTc0ODUxMDY2MiwiaXNzIjoi5ZSQ5L-K5ZacIiwicm9sZV9pZCI6IlI1NzIwMzc3NzcyMjE4NzgwOCIsInN1YiI6ImxvZ2luIiwidXNlcl9pZCI6IjQyNiIsIndhcmVob3VzZV9pZHMiOnsiVzU3MjAzNzcwMzA5NzM4NDk2IjoiVzU3MjAzNzcwMzA5NzM4NDk2In19.MJuVPpnrYaGFygsiaE--peOwNmjEUC16plr4d0TN2qY"

# 请求头
headers = {
    "authority": "devops-api.nullmax.net",
    "accept": "application/json, text/plain, */*",
    "accept-encoding": "gzip, deflate, br",
    "accept-language": "en-US,en;q=0.9",
    "authorization": AUTH_TOKEN,
    "content-type": "application/json",
    "origin": "https://devops.nullmax.net",
    "referer": "https://devops.nullmax.net/",
    "sec-ch-ua": '"Google Chrome";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Linux"',
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "same-site",
    "user-agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.54 Safari/537.36",
    "warehouse": "W57203770309738496"
}

# 请求体数据
payload = {
    "id": "",
    "employeeId": "",
    "warehouseId": "",
    "remark": ""
}

def transfer_asset() -> Optional[Dict]:
    """
    执行固定资产调拨/重新分配操作
    :return: 包含响应数据的字典，出错时返回None
    """
    try:
        response = requests.put(
            API_URL,
            headers=headers,
            data=json.dumps(payload),
            timeout=10
        )
        
        # 检查HTTP状态码
        response.raise_for_status()
        
        # 解析JSON响应
        result = response.json()
        
        # 验证响应结构
        if not isinstance(result, dict):
            raise ValueError("响应不是有效的JSON对象")
            
        return result
        
    except requests.exceptions.RequestException as e:
        print(f"🔴 请求失败: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"错误响应: {e.response.text}")
        return None
    except ValueError as e:
        print(f"🔴 JSON解析错误: {e}")
        #print(f"原始响应: {response.text}")
        return None

def display_result(result: Dict) -> None:
    """格式化显示API响应结果"""
    if not result:
        return
    
    print("\n✅ 资产调拨操作结果:")
    print(f"• 状态码: {result.get('code', 'N/A')}")
    print(f"• 消息: {result.get('msg', '无消息')}")
    
    if 'data' in result:
        print("• 返回数据:")
        for key, value in result['data'].items():
            print(f"  {key}: {value}")


import json
import openpyxl
from openpyxl.utils import column_index_from_string

def read_json_file(json_file):
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data

def process_excel(json_data, excel_file):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    
    # 获取资产映射
    asset_list = json_data['data']['list']
    asset_map = {
                  asset['assetCode']: {
                                        'id': asset['id'], 
                                        'warehouseId': asset['warehouseId']
                                    } 
                  for asset in asset_list
                    if 'assetCode' in asset and str(asset['assetCode']).upper().startswith("BJB")
                }
    print(asset_map)
    # 确定列位置（基于你的表头）
    headers = [cell.value for cell in sheet[1]]
    print(headers)

    # 编号列位置（第一列）
    code_col = column_index_from_string('A')
    
    # 资产ID和仓库ID列位置
    id_col = None
    warehouse_col = None
    
    # 检查列是否已存在
    for col_idx, header in enumerate(headers, start=1):
        if header == "资产ID":
            id_col = col_idx
        elif header == "仓库ID":
            warehouse_col = col_idx

     # 如果列不存在，添加到最后一列之后
    last_col = len(headers)
    if id_col is None:
        id_col = last_col + 1
        sheet.cell(row=1, column=id_col, value="资产ID")
    if warehouse_col is None:
        warehouse_col = last_col + 2
        sheet.cell(row=1, column=warehouse_col, value="仓库ID")
        
    # 遍历Excel行（从第2行开始）
    for row in range(2, sheet.max_row + 1):
        asset_code = sheet.cell(row=row, column=code_col).value
        #print(asset_code)
        if asset_code:  # 确保单元格不为空
            # 将Excel中的编号转换为大写进行比较
            asset_code = str(asset_code).upper()
            print(asset_code)
            if asset_code in asset_map:
                sheet.cell(row=row, column=id_col, value=asset_map[asset_code]['id'])
                sheet.cell(row=row, column=warehouse_col, value=asset_map[asset_code]['warehouseId'])

    # 保存Excel文件
    wb.save(excel_file)
    print(f"Excel文件已更新并保存: {excel_file}")
    

def main():
    json_file = '/data/myproject/insert-asserts/fixed_assets.json'  # 替换为你的JSON文件路径
    excel_file = '/home/ubuntu/Downloads/output6_backup.xlsx'  # 替换为你的Excel文件路径
    
    json_data = read_json_file(json_file)
    process_excel(json_data, excel_file)

if __name__ == '__main__':
    main()


# if __name__ == "__main__":
#     print("🔄 正在执行固定资产调拨操作...")
    
#     # 执行调拨操作
#     response_data = transfer_asset()